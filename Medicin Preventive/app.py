import base64
from flask import Flask, request, render_template  # type: ignore
from src.data.disease_info import DiseaseInfo
from src.models.dp_model import DiseasePredictModel
from googletrans import Translator  # type: ignore
import src.visuals.piechart as pie
import pickle
import numpy as np # type: ignore
import pandas as pd # type: ignore
from openpyxl import Workbook, load_workbook # type: ignore

app = Flask(__name__)

# Charger les modèles et scaler
scaler = pickle.load(open('models/scaler.pkl', 'rb'))
lr = pickle.load(open('models/lr.pkl', 'rb'))

# Initialisation des classes et du traducteur
di = DiseaseInfo()
dpm = DiseasePredictModel()
translator = Translator()

# Fonction pour ajouter une prédiction dans un fichier Excel
def ajouter_prediction_excel(fichier, donnees):
    try:
        # Charger le fichier s'il existe, sinon en créer un nouveau
        try:
            workbook = load_workbook(fichier)
            feuille = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            feuille = workbook.active
            feuille.append(['Glucose', 'Age', 'BloodPressure', 'Insulin', 'BMI', 'SkinThickness', 'DiabetesPedigreeFunction', 'Prediction'])

        # Ajouter les données à la feuille
        feuille.append([donnees['Glucose'], donnees['Age'], donnees['BloodPressure'], donnees['Insulin'], 
                        donnees['BMI'], donnees['SkinThickness'], donnees['DiabetesPedigreeFunction'], donnees['Prediction']])

        # Enregistrer le fichier
        workbook.save(fichier)

    except Exception as e:
        print("Erreur lors de l'ajout dans le fichier Excel :", e)


# Fonction de traduction
def translate_to_french(text: str) -> str:
    """Translate the given text to French using Google Translate."""
    try:
        translated = translator.translate(text, src='en', dest='fr')
        return translated.text
    except Exception as e:
        print(f"(dinfo-module)[Error]: Translation failed - {e}")
        return text  # Retourner le texte original en cas d'échec

@app.route("/diseases")
def disease_page():
    fr_symp = []
    
    # Vérifie si les symptômes sont disponibles
    if dpm.symptoms is None:
        return render_template('500.html'), 500
    
    # Traduction des symptômes en français
    for sym in dpm.symptoms:
        trans = translate_to_french(sym)
        fr_symp.append(trans)
    
    return render_template('index.html', symp_names=fr_symp)

@app.route("/")
def index() :
    return render_template("disease.html")

@app.route("/predict", methods=['GET'])
def predict():
    """Prediction des maladies à travers les symptômes."""
    user_symp = request.args.get('symptoms', 'unknown').split(',')
    pred_probs = dpm.predict_proba(user_symp)

    if pred_probs is None or dpm.diseases is None:
        return render_template('500.html'), 500

    labels, probs = pie.filter_proba(dpm.diseases, pred_probs, 0.05)
    pred_disease = labels[max(range(len(probs)), key=lambda i: probs[i])]
    
    pie_image = pie.disease_piechart(labels, probs)

    return render_template(
        'result.html',
        d_name=pred_disease,
        d_summary=di.short_summary(pred_disease),
        d_image_url=di.image_link(pred_disease),
        d_precautions=di.precautions(pred_disease),
        piechart_data=base64.b64encode(pie_image.read()).decode()
    )


@app.route('/resultdisease', methods=['POST'])
def resultdisease():
    # Récupérer les valeurs du formulaire
    Age = int(request.form.get("Age"))
    Glucose = int(request.form.get("Glucose"))
    BloodPressure = int(request.form.get("BloodPressure"))
    Insulin = int(request.form.get("Insulin"))
    BMIs = float(request.form.get("BMI"))
    SkinThickness = int(request.form.get("SkinThickness"))
    DiabetesPedigreeFunction = float(request.form.get("DiabetesPedigreeFunction"))

    # Préparer les données pour la prédiction
    temp_arr = [Glucose, BloodPressure, SkinThickness, Insulin, BMIs, DiabetesPedigreeFunction, Age]
    data = np.array([temp_arr])
    temp_sc = scaler.transform(data)
    pred = lr.predict(temp_sc)[0]

    # Préparer les données pour Excel
    donnees_personne = {
        'Glucose': Glucose,
        'Age': Age,
        'BloodPressure': BloodPressure,
        'Insulin': Insulin,
        'BMI': BMIs,
        'SkinThickness': SkinThickness,
        'DiabetesPedigreeFunction': DiabetesPedigreeFunction,
        'Prediction': int(pred)
    }

    # Enregistrer dans le fichier Excel
    ajouter_prediction_excel("predictions_diabetes.xlsx", donnees_personne)

    # Déterminer le texte du résultat
    if pred == 0:
        res = "does not indicate diabetes"
    else:
        res = "indicates diabetes"

    return render_template('resultdisease.html', prediction=res)

if __name__ == "__main__":
    app.run(debug=True)
